import oracledb
import pandas as pd
from openpyxl import load_workbook
import os
from datetime import datetime

# Obter a data atual
data_atual = datetime.now()

# Extrair o ano e o mês e formatar no formato AAAAMM
referencia = data_atual.strftime("%Y%m")

referencia_arquivo = data_atual.strftime("%m-%Y")

print(referencia)
""" referencia = input("Digite a referecia: ") """



# Criação da pasta 'relatorios' se ela não existir
output_dir = 'relatorios'
if not os.path.exists(output_dir):
    os.makedirs(output_dir)
    

# Configurações de conexão para cada banco de dados
db_configs = [
    {
        'name': 'DESO',
        'user': 'NETUNO_CONSULTA',
        'password': 'oDDanqeMNpwzaCSG',
        'dsn': '10.41.36.2/PDBEOS',
        'sql': f"""
                    SELECT C.ANO_MES_LEITURA "REFERÊNCIA", C.ID_CICLO CICLO, C.ROTA, C.PAGINA, TRUNC(C.DATA_LEITURA) DATA_LEITURA, C.MATRICULA, 
            '9-Enviado para simultanea' AS MODO_FAT, 
            CASE C.ID_CONF 
                    WHEN 0 THEN '0-Leitura nao Coletada'
                    WHEN 1 THEN '1-Leitura nao Informada' 
                    WHEN 2 THEN '2-Leitura Fora da Faixa nao confirmada'
                    WHEN 3 THEN '3-Leitura Normal'    
                    WHEN 4 THEN '4-Leitura Fora de Faixa Confirmada'   
                    WHEN 5 THEN '5-Leitura Retificada Dentro da Faixa' 
                    WHEN 6 THEN '6-Leitura Retificada Fora de Faixa' 
                    WHEN 9 THEN '9-Nao Medido'
                    ELSE 'DESCONHECIDO'
            END ID_CONF,       
            CASE C.SIT_AGUA 
                    WHEN 1 THEN 'ATIVO' 
                    WHEN 2 THEN 'CORTADO POR DEBITO' 
                    WHEN 3 THEN 'CORTADO A PEDIDO' 
                    WHEN 4 THEN 'SUPRIMIDO' 
                    WHEN 5 THEN 'FACTIVEL' 
                    WHEN 6 THEN 'POTENCIAL' 
                    WHEN 9 THEN 'EXCLUIDO' 
                    ELSE 'DESCONHECIDO' 
            END SITUACAO_AGUA,
            CASE C.SIT_ESGOTO 
                    WHEN 1 THEN 'ATIVO' 
                    WHEN 2 THEN 'CORTADO POR DEBITO' 
                    WHEN 3 THEN 'CORTADO A PEDIDO' 
                    WHEN 4 THEN 'SUPRIMIDO' 
                    WHEN 5 THEN 'FACTIVEL' 
                    WHEN 6 THEN 'POTENCIAL' 
                    WHEN 9 THEN 'EXCLUIDO' 
                    ELSE 'DESCONHECIDO' 
            END SITUACAO_ESGOTO,
            CA.INDICEA, 
            CASE CA.SITUACAO 
                    WHEN 0 THEN '0-Gerado'
                    WHEN 1 THEN '1-Exportado' 
                    WHEN 2 THEN '2-Importado'
                    WHEN 3 THEN '3-Processado com sucesso'    
                    WHEN 4 THEN '4-Programado'   
                    WHEN 5 THEN '5-Processado com erro'
                    ELSE 'DESCONHECIDO'
            END SITUACAO_CA       
        FROM PRD_DESO_NETUNO.CONSUMO C
        LEFT JOIN PRD_DESO_NETUNO.CONTROLE_ARQUIVO CA 
            ON CA.ANO_MES = C.ANO_MES_LEITURA 
            AND CA.ID_CICLO = C.ID_CICLO 
            AND CA.ROTA = C.ROTA 
            AND CA.PAGINA = C.PAGINA 
        WHERE C.MODO_FAT='9' 
        AND C.ANO_MES_LEITURA>= '{int(referencia) - 1}'
        AND (C.SIT_AGUA<4 OR C.SIT_ESGOTO=1) 
        AND TRUNC(C.DATA_LEITURA) < TRUNC(SYSDATE) 
        AND C.PAGINA>0
        ORDER BY C.ANO_MES_LEITURA DESC, C.ID_CICLO, C.ROTA, C.PAGINA
        """,
 'output': os.path.join(output_dir, f'SDES-2802_Relatório_ligações_não_atualizadas_na_LIS_({referencia_arquivo}).xlsx'),
        'template': f'./modelos/SDES-2802_Relatório_ligações_não_atualizadas_na_LIS.xlsx'
    },
    {
            'name': 'ATS',
            'user': 'PRD_ATS_NETUNO',
            'password': '0bKRXWgG9QVElmzA16ru',
            'dsn': '10.41.36.2/PDBEOS',
            'sql': f"""
            SELECT C.ANO_MES_LEITURA "REFERÊNCIA", C.ID_CICLO CICLO, C.ROTA, C.PAGINA, TRUNC(C.DATA_LEITURA) DATA_LEITURA, C.MATRICULA, 
        '9-Enviado para simultanea' AS MODO_FAT, 
        CASE C.ID_CONF 
                WHEN 0 THEN '0-Leitura nao Coletada'
                WHEN 1 THEN '1-Leitura nao Informada' 
                WHEN 2 THEN '2-Leitura Fora da Faixa nao confirmada'
                WHEN 3 THEN '3-Leitura Normal'    
                WHEN 4 THEN '4-Leitura Fora de Faixa Confirmada'   
                WHEN 5 THEN '5-Leitura Retificada Dentro da Faixa' 
                WHEN 6 THEN '6-Leitura Retificada Fora de Faixa' 
                WHEN 9 THEN '9-Nao Medido'
                ELSE 'DESCONHECIDO'
        END ID_CONF,       
        CASE C.SIT_AGUA 
                WHEN 1 THEN 'ATIVO' 
                WHEN 2 THEN 'CORTADO POR DEBITO' 
                WHEN 3 THEN 'CORTADO A PEDIDO' 
                WHEN 4 THEN 'SUPRIMIDO' 
                WHEN 5 THEN 'FACTIVEL' 
                WHEN 6 THEN 'POTENCIAL' 
                WHEN 9 THEN 'EXCLUIDO' 
                ELSE 'DESCONHECIDO' 
        END SITUACAO_AGUA,
        CASE C.SIT_ESGOTO 
                WHEN 1 THEN 'ATIVO' 
                WHEN 2 THEN 'CORTADO POR DEBITO' 
                WHEN 3 THEN 'CORTADO A PEDIDO' 
                WHEN 4 THEN 'SUPRIMIDO' 
                WHEN 5 THEN 'FACTIVEL' 
                WHEN 6 THEN 'POTENCIAL' 
                WHEN 9 THEN 'EXCLUIDO' 
                ELSE 'DESCONHECIDO' 
        END SITUACAO_ESGOTO,
        CA.INDICEA, 
        CASE CA.SITUACAO 
                WHEN 0 THEN '0-Gerado'
                WHEN 1 THEN '1-Exportado' 
                WHEN 2 THEN '2-Importado'
                WHEN 3 THEN '3-Processado com sucesso'    
                WHEN 4 THEN '4-Programado'   
                WHEN 5 THEN '5-Processado com erro'
                ELSE 'DESCONHECIDO'
        END SITUACAO_CA       
    FROM PRD_ATS_NETUNO.CONSUMO C
    LEFT JOIN PRD_ATS_NETUNO.CONTROLE_ARQUIVO CA 
        ON CA.ANO_MES = C.ANO_MES_LEITURA 
        AND CA.ID_CICLO = C.ID_CICLO 
        AND CA.ROTA = C.ROTA 
        AND CA.PAGINA = C.PAGINA 
    WHERE C.MODO_FAT='9' 
    AND C.ANO_MES_LEITURA>='{int(referencia) - 1}'
    AND (C.SIT_AGUA<4 OR C.SIT_ESGOTO=1) 
    AND TRUNC(C.DATA_LEITURA) < TRUNC(SYSDATE) 
    AND C.PAGINA>0
    ORDER BY C.ANO_MES_LEITURA DESC, C.ID_CICLO, C.ROTA, C.PAGINA
        """,
        'output': os.path.join(output_dir, f'ATS-4825_Relatório_ligações_não_atualizadas_na_LIS_({referencia_arquivo}).xlsx'),
        'template': f'./modelos/ATS-4825_Relatório_ligações_não_atualizadas_na_LIS.xlsx'
    },
    {
        'name': 'CODAU',
        'user': 'CODAU_CONSULTA',
        'password': 'q6eMDXqbuQIUKuQI',
        'dsn': '10.41.36.2/PDBEOS',
        'sql': f"""
                    SELECT C.ANO_MES_LEITURA "REFERÊNCIA", C.ID_CICLO CICLO, C.ROTA, C.PAGINA, TRUNC(C.DATA_LEITURA) DATA_LEITURA, C.MATRICULA, 
            '9-Enviado para simultanea' AS MODO_FAT, 
            CASE C.ID_CONF 
                    WHEN 0 THEN '0-Leitura nao Coletada'
                    WHEN 1 THEN '1-Leitura nao Informada' 
                    WHEN 2 THEN '2-Leitura Fora da Faixa nao confirmada'
                    WHEN 3 THEN '3-Leitura Normal'    
                    WHEN 4 THEN '4-Leitura Fora de Faixa Confirmada'   
                    WHEN 5 THEN '5-Leitura Retificada Dentro da Faixa' 
                    WHEN 6 THEN '6-Leitura Retificada Fora de Faixa' 
                    WHEN 9 THEN '9-Nao Medido'
                    ELSE 'DESCONHECIDO'
            END ID_CONF,       
            CASE C.SIT_AGUA 
                    WHEN 1 THEN 'ATIVO' 
                    WHEN 2 THEN 'CORTADO POR DEBITO' 
                    WHEN 3 THEN 'CORTADO A PEDIDO' 
                    WHEN 4 THEN 'SUPRIMIDO' 
                    WHEN 5 THEN 'FACTIVEL' 
                    WHEN 6 THEN 'POTENCIAL' 
                    WHEN 9 THEN 'EXCLUIDO' 
                    ELSE 'DESCONHECIDO' 
            END SITUACAO_AGUA,
            CASE C.SIT_ESGOTO 
                    WHEN 1 THEN 'ATIVO' 
                    WHEN 2 THEN 'CORTADO POR DEBITO' 
                    WHEN 3 THEN 'CORTADO A PEDIDO' 
                    WHEN 4 THEN 'SUPRIMIDO' 
                    WHEN 5 THEN 'FACTIVEL' 
                    WHEN 6 THEN 'POTENCIAL' 
                    WHEN 9 THEN 'EXCLUIDO' 
                    ELSE 'DESCONHECIDO' 
            END SITUACAO_ESGOTO,
            CA.INDICEA, 
            CASE CA.SITUACAO 
                    WHEN 0 THEN '0-Gerado'
                    WHEN 1 THEN '1-Exportado' 
                    WHEN 2 THEN '2-Importado'
                    WHEN 3 THEN '3-Processado com sucesso'    
                    WHEN 4 THEN '4-Programado'   
                    WHEN 5 THEN '5-Processado com erro'
                    ELSE 'DESCONHECIDO'
            END SITUACAO_CA       
        FROM PRD_CODAU_NETUNO.CONSUMO C
        LEFT JOIN PRD_CODAU_NETUNO.CONTROLE_ARQUIVO CA 
            ON CA.ANO_MES = C.ANO_MES_LEITURA 
            AND CA.ID_CICLO = C.ID_CICLO 
            AND CA.ROTA = C.ROTA 
            AND CA.PAGINA = C.PAGINA 
        WHERE C.MODO_FAT='9' 
        AND C.ANO_MES_LEITURA>='{int(referencia) - 1}'
        AND (C.SIT_AGUA<4 OR C.SIT_ESGOTO=1) 
        AND TRUNC(C.DATA_LEITURA) < TRUNC(SYSDATE) 
        AND C.PAGINA>0
        ORDER BY C.ANO_MES_LEITURA DESC, C.ID_CICLO, C.ROTA, C.PAGINA
        """,
        'output': os.path.join(output_dir, f'SCOD-394_Relatório_ligações_não_atualizadas_na_LIS_({referencia_arquivo}).xlsx'),
        'template': f'./modelos/SCOD-394_Relatório_ligações_não_atualizadas_na_LIS.xlsx'
    }
]

# Função para inserir os dados no modelo de relatório
def insert_data_into_template(df, template_path, output_path):
    # Carregar o modelo
    wb = load_workbook(template_path)
    ws = wb.active
    
      # Formatar as datas no DataFrame para DD/MM/YYYY
    df['DATA_LEITURA'] = df['DATA_LEITURA'].dt.strftime('%d/%m/%Y')
    
    # Inserir os dados a partir da linha 5
    for r_idx, row in df.iterrows():
        for c_idx, value in enumerate(row):
            ws.cell(row=r_idx + 5, column=c_idx + 1, value=value)
    
     # Inserir a data atual na coluna L, linha 2
    data_atual = datetime.now().strftime('%d/%m/%Y')
    ws.cell(row=2, column=12, value=data_atual)  # Coluna L é a coluna 12 em indexação de célula (1-based index)
    # Salvar o arquivo de saída
    wb.save(output_path)

# Executar a consulta para cada banco de dados e salvar os resultados nos modelos de relatórios
for db_config in db_configs:
    try:
        # Conectar ao banco de dados
        connection = oracledb.connect(
            user=db_config['user'],
            password=db_config['password'],
            dsn=db_config['dsn']
        )
        cursor = connection.cursor()
        
        # Executar a consulta
        cursor.execute(db_config['sql'])
        
        # Recuperar os resultados
        columns = [col[0] for col in cursor.description]
        results = cursor.fetchall()
        
        # Verificar se há resultados
        if not results:
            print(f"Nenhum resultado encontrado para o banco de dados {db_config['name']}. Arquivo Excel não será criado.")
            continue
        
        # Converter os resultados em um DataFrame do pandas
        df = pd.DataFrame(results, columns=columns)
        
        # Inserir os dados no modelo de relatório e salvar o arquivo
        insert_data_into_template(df, db_config['template'], db_config['output'])
        print(f"Resultados do banco de dados {db_config['name']} salvos em {db_config['output']}")

    except oracledb.Error as e:
        print(f"Erro ao executar a consulta no banco de dados {db_config['name']}:", e)

    finally:
        # Fechar cursor e conexão
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals():
            connection.close()
     
input("Pressione Enter para sair...")
            
            
